#!/usr/bin/env python3
"""
Import Qatar PC Shift Table .xlsx → Supabase (stores_qa, promoters_qa, shifts_qa)
"""

import json
import os
import re
import ssl
import urllib.request
import urllib.error
from datetime import datetime

import openpyxl

# Fix macOS Python SSL cert issue
ssl._create_default_https_context = ssl._create_unverified_context

# --- Config ---
# Set SUPABASE_SERVICE_ROLE_KEY env var before running
XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '🇶🇦Qatar PC Shift Table .xlsx')
SUPABASE_URL = os.environ.get('SUPABASE_URL', 'https://bsbxvntzddqwjhdnjvof.supabase.co')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '')

SPECIAL_SHIFTS = {'Off', 'AL', 'SL', 'LOP'}


def supabase_request(method, path, data=None):
    url = SUPABASE_URL + path
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation',
    }
    if method == 'POST' and 'on_conflict' in path:
        headers['Prefer'] = 'resolution=merge-duplicates,return=representation'

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        print(f'  ERROR {e.code}: {e.read().decode()}')
        raise


def parse_time(t):
    """Convert '13.00' or '13:00' → '13:00'"""
    if not t:
        return None
    s = str(t).strip().replace('.', ':')
    # handle '24:00' → keep as-is (or '00:00' next day, but Supabase TIME accepts it)
    m = re.match(r'^(\d{1,2}):(\d{2})$', s)
    if m:
        return f'{int(m.group(1)):02d}:{m.group(2)}'
    return None


def main():
    print(f'Loading {XLSX_PATH}...')
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # ========== 1. STORES ==========
    print('\n=== Importing Stores → stores_qa ===')
    ws_store = wb['Store setting']
    stores = {}  # abbr → {name, open_time, close_time, shift_slots, active}

    for r in range(2, ws_store.max_row + 1):
        store_name = ws_store.cell(r, 1).value
        abbr = ws_store.cell(r, 2).value
        shift = ws_store.cell(r, 3).value
        active = ws_store.cell(r, 7).value

        if not abbr or not store_name:
            continue
        abbr = str(abbr).strip()
        if abbr in SPECIAL_SHIFTS:
            continue

        if abbr not in stores:
            stores[abbr] = {
                'code': abbr,
                'name': str(store_name).strip(),
                'active': active in (True, 'True', 'TRUE'),
                'shift_slots': [],
            }

        if shift:
            shift_str = str(shift).strip()
            # Parse "13.00-22.00" → open/close times
            m = re.match(r'^(\d{1,2}[.:]\d{2})\s*-\s*(\d{1,2}[.:]\d{2})$', shift_str)
            if m:
                open_t = parse_time(m.group(1))
                close_t = parse_time(m.group(2))
                slot = f'{open_t}-{close_t}'
                if slot not in stores[abbr]['shift_slots']:
                    stores[abbr]['shift_slots'].append(slot)

    # Set open/close from first shift slot
    store_rows = []
    for code, info in stores.items():
        slots = info['shift_slots']
        open_time = '10:00'
        close_time = '22:00'
        if slots:
            first = slots[0].split('-')
            open_time = first[0]
            close_time = first[1]

        store_rows.append({
            'code': code,
            'name': info['name'],
            'active': info['active'],
            'open_time': open_time,
            'close_time': close_time,
            'shift_slots': slots if slots else None,
        })

    print(f'  Found {len(store_rows)} stores')
    for s in store_rows:
        print(f'    {s["code"]}: {s["name"]} ({s["open_time"]}-{s["close_time"]}) slots={s["shift_slots"]}')

    result = supabase_request('POST', '/rest/v1/stores_qa?on_conflict=code', store_rows)
    store_id_map = {}  # code → uuid
    for row in result:
        store_id_map[row['code']] = row['id']
    print(f'  ✅ Upserted {len(result)} stores')

    # ========== 2. PROMOTERS ==========
    print('\n=== Importing Promoters → promoters_qa ===')
    ws_pc = wb['PC setting']
    promoter_rows = []

    for r in range(2, ws_pc.max_row + 1):
        code = ws_pc.cell(r, 2).value  # Code (e.g. "Ahmed Ha")
        if not code:
            continue
        code = str(code).strip()
        nickname = ws_pc.cell(r, 4).value  # ชื่อเล่น
        first_name = ws_pc.cell(r, 5).value
        last_name = ws_pc.cell(r, 6).value

        # Build full name for display
        full_name = code  # fallback
        if first_name and last_name:
            full_name = f'{str(first_name).strip()} {str(last_name).strip()}'
        elif first_name:
            full_name = str(first_name).strip()

        promoter_rows.append({
            'name': code,  # use short code as name (matches Trans table headers)
            'stores_label': full_name,  # store full name in stores_label for display
            'active': True,
        })

    print(f'  Found {len(promoter_rows)} promoters')
    result = supabase_request('POST', '/rest/v1/promoters_qa?on_conflict=name', promoter_rows)
    promoter_id_map = {}  # name → uuid
    for row in result:
        promoter_id_map[row['name']] = row['id']
    print(f'  ✅ Upserted {len(result)} promoters')

    # ========== 3. SHIFTS ==========
    print('\n=== Importing Shifts → shifts_qa ===')
    ws_trans = wb['Trans table']

    # Header row: [Shift table, Ahmed Ha, Akram Be, ...]
    promoter_names = []
    for c in range(2, ws_trans.max_column + 1):
        val = ws_trans.cell(1, c).value
        if val:
            promoter_names.append((c, str(val).strip()))

    print(f'  Promoters in Trans table: {len(promoter_names)}')

    shift_rows = []
    skipped = 0

    for r in range(2, ws_trans.max_row + 1):
        date_val = ws_trans.cell(r, 1).value
        if not date_val:
            continue

        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            try:
                date_str = datetime.strptime(str(date_val), '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
            except ValueError:
                continue

        for col, pname in promoter_names:
            cell = ws_trans.cell(r, col).value
            if not cell:
                continue
            cell_str = str(cell).strip()
            if not cell_str or cell_str == '-':
                continue

            promoter_id = promoter_id_map.get(pname)
            if not promoter_id:
                skipped += 1
                continue

            # Parse cell: "VLM 13.00-22.00" or "Off" or "SL" etc.
            shift_type = cell_str
            time_range = None

            parts = cell_str.split()
            if len(parts) >= 2:
                store_code = parts[0]
                time_part = parts[1]
                m = re.match(r'^(\d{1,2}[.:]\d{2})\s*-\s*(\d{1,2}[.:]\d{2})$', time_part)
                if m:
                    shift_type = store_code
                    t1 = parse_time(m.group(1))
                    t2 = parse_time(m.group(2))
                    time_range = f'{t1}-{t2}'
            elif cell_str in SPECIAL_SHIFTS:
                shift_type = cell_str
                time_range = None

            shift_rows.append({
                'promoter_id': promoter_id,
                'date': date_str,
                'shift_type': shift_type,
                'time_range': time_range,
            })

    print(f'  Total shifts to upsert: {len(shift_rows)} (skipped {skipped} unmatched)')

    # Upsert in batches
    batch_size = 500
    for i in range(0, len(shift_rows), batch_size):
        batch = shift_rows[i:i + batch_size]
        supabase_request('POST', '/rest/v1/shifts_qa?on_conflict=promoter_id,date', batch)
        print(f'  Batch {i // batch_size + 1}: {len(batch)} shifts upserted')

    print(f'\n✅ Qatar import complete!')
    print(f'   Stores: {len(store_rows)}')
    print(f'   Promoters: {len(promoter_rows)}')
    print(f'   Shifts: {len(shift_rows)}')


if __name__ == '__main__':
    main()
